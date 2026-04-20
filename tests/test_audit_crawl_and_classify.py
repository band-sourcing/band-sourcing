"""
[e2e] audit_crawl_and_classify.py 테스트.

크롤링 부분은 mock으로 대체하고, 파싱/분류/엑셀출력/통계 부분을 실제 동작으로 검증.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.audit_crawl_and_classify import (  # noqa: E402
    EXCEL_COLUMNS,
    build_row_from_parse_error,
    build_row_from_product,
    build_stats,
    write_excel,
)
from src.content_parser import parse_post  # noqa: E402


# ──────────────────────────────────────────────────────────────
# 테스트 픽스처
# ──────────────────────────────────────────────────────────────
@pytest.fixture
def brand_map():
    return {
        "#FG": "페레가모",
        "#GC": "구찌",
        "#NK": "나이키",
        "#RL": "로렉스",
        "#AZ": "어메이징크리",
    }


@pytest.fixture
def category_keywords():
    return {
        "bag": ["가방", "백", "토트", "숄더백"],
        "watch": ["시계", "워치"],
        "wallet": ["지갑", "카드홀더"],
        "shoes": ["스니커즈", "로퍼"],
        "outer": ["자켓", "패딩", "코트"],
        "top": ["반팔", "티셔츠", "후디", "맨투맨"],
        "bottom": ["데님", "슬랙스", "팬츠"],
        "accessory": ["벨트", "모자", "선글라스"],
    }


@pytest.fixture
def exclusion_config():
    return {
        "factory_codes": {
            "enabled": True,
            "codes": ["qe", "bs", "af"],
        },
        "free_size": {
            "enabled": True,
            "target_bands": ["의류천국22"],
        },
    }


# ──────────────────────────────────────────────────────────────
# build_row_from_product 테스트
# ──────────────────────────────────────────────────────────────
class TestBuildRowFromProduct:
    def test_normal_product_classified_correctly(
        self, brand_map, category_keywords, exclusion_config
    ):
        content = """#FG
반팔 티셔츠
사이즈-M L XL
028 (QT)"""
        products = parse_post(content, brand_map, "의류천국22")
        assert len(products) == 1

        row = build_row_from_product(
            product=products[0],
            post_key="post_123",
            source_band="의류천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )

        assert row["post_key"] == "post_123"
        assert row["source_band"] == "의류천국22"
        assert row["brand_tag"] == "#FG"
        assert row["brand_name_kr"] == "페레가모"
        assert row["auto_category"] == "top"
        assert row["price_code"] == "QT"
        assert row["cost_price"] == 28000
        assert row["parse_status"] == "ok"
        assert row["excluded_reason"] == ""
        assert row["correct_category"] == ""
        assert row["txtBody_raw"] == content

    def test_excluded_by_factory_code_qe(
        self, brand_map, category_keywords, exclusion_config
    ):
        content = """#GC
토트백
050 (qe)"""
        products = parse_post(content, brand_map, "잡화천국22")
        row = build_row_from_product(
            product=products[0],
            post_key="post_qe",
            source_band="잡화천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )
        # classify_category는 여전히 동작하되 제외 사유 기록됨
        assert row["auto_category"] == "bag"
        assert "factory_code" in row["excluded_reason"]

    def test_excluded_by_free_size_in_target_band(
        self, brand_map, category_keywords, exclusion_config
    ):
        content = """#NK
후디
사이즈-FREE
045 (BM)"""
        products = parse_post(content, brand_map, "의류천국22")
        row = build_row_from_product(
            product=products[0],
            post_key="post_free",
            source_band="의류천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )
        assert row["excluded_reason"] == "free_size"

    def test_free_size_not_excluded_in_non_target_band(
        self, brand_map, category_keywords, exclusion_config
    ):
        # 잡화천국22는 free_size exclusion target이 아님
        content = """#GC
벨트
사이즈-FREE
045 (QT)"""
        products = parse_post(content, brand_map, "잡화천국22")
        row = build_row_from_product(
            product=products[0],
            post_key="post_bag_free",
            source_band="잡화천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )
        assert row["excluded_reason"] == ""
        assert row["auto_category"] == "accessory"

    def test_watch_brand_fallback(
        self, brand_map, category_keywords, exclusion_config
    ):
        # RL(로렉스) 브랜드는 키워드 없어도 watch로 분류됨 (margin_engine fallback)
        content = """#RL
데이저스트
120 (QT)"""
        products = parse_post(content, brand_map, "잡화천국22")
        row = build_row_from_product(
            product=products[0],
            post_key="post_watch",
            source_band="잡화천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )
        # "데이저스트"는 위 category_keywords에 없으므로 fallback으로 watch
        assert row["auto_category"] == "watch"

    def test_etc_when_no_keyword_match(
        self, brand_map, category_keywords, exclusion_config
    ):
        # 키워드 미매칭 + 시계브랜드 아님 → etc
        content = """#GC
알수없는상품명ZZZ
050 (QT)"""
        products = parse_post(content, brand_map, "잡화천국22")
        row = build_row_from_product(
            product=products[0],
            post_key="post_etc",
            source_band="잡화천국22",
            brand_map=brand_map,
            category_keywords=category_keywords,
            exclusion_config=exclusion_config,
            txtbody_raw=content,
        )
        assert row["auto_category"] == "etc"

    def test_set_product_generates_suffixed_post_keys(
        self, brand_map, category_keywords, exclusion_config
    ):
        content = """#AZ
"네오테크 후디 셋업"
색상: 그레이
상의: 95 100 105
하의: 30 32 34
상의 053 (AL)
하의 046 (AL)"""
        products = parse_post(content, brand_map, "의류천국22")
        assert len(products) == 2

        rows = [
            build_row_from_product(
                product=p,
                post_key="post_set",
                source_band="의류천국22",
                brand_map=brand_map,
                category_keywords=category_keywords,
                exclusion_config=exclusion_config,
                txtbody_raw=content,
            )
            for p in products
        ]

        post_keys = [r["post_key"] for r in rows]
        assert "post_set_top" in post_keys
        assert "post_set_bottom" in post_keys


# ──────────────────────────────────────────────────────────────
# build_row_from_parse_error 테스트
# ──────────────────────────────────────────────────────────────
class TestBuildRowFromParseError:
    def test_parse_error_row_has_correct_status(self):
        row = build_row_from_parse_error(
            post_key="post_err",
            source_band="잡화천국22",
            txtbody_raw="브랜드 없는 게시글",
            error_msg="브랜드 태그를 찾을 수 없음",
        )
        assert row["parse_status"] == "parse_error"
        assert row["auto_category"] == "etc"
        assert "PARSE_ERROR" in row["product_name"]
        assert row["brand_tag"] == ""
        assert row["txtBody_raw"] == "브랜드 없는 게시글"


# ──────────────────────────────────────────────────────────────
# write_excel 테스트
# ──────────────────────────────────────────────────────────────
class TestWriteExcel:
    def test_excel_output_structure(self, tmp_path):
        rows = [
            {
                "post_key": "post_1",
                "source_band": "잡화천국22",
                "brand_tag": "#GC",
                "brand_name_kr": "구찌",
                "product_name": "토트백",
                "price_code": "QT",
                "cost_price": 250000,
                "sizes": "M, L",
                "auto_category": "bag",
                "excluded_reason": "",
                "parse_status": "ok",
                "correct_category": "",
                "txtBody_raw": "원본 텍스트",
            },
            {
                "post_key": "post_2",
                "source_band": "의류천국22",
                "brand_tag": "#FG",
                "brand_name_kr": "페레가모",
                "product_name": "반팔",
                "price_code": "QT",
                "cost_price": 28000,
                "sizes": "M, L, XL",
                "auto_category": "top",
                "excluded_reason": "",
                "parse_status": "ok",
                "correct_category": "",
                "txtBody_raw": "반팔 원본",
            },
        ]

        output = tmp_path / "audit_test.xlsx"
        write_excel(rows, output)

        assert output.exists()

        wb = load_workbook(output)
        ws = wb.active

        # 헤더 검증
        header_row = [ws.cell(row=1, column=i + 1).value for i in range(len(EXCEL_COLUMNS))]
        expected_headers = [c[0] for c in EXCEL_COLUMNS]
        assert header_row == expected_headers

        # 데이터 row 수
        assert ws.max_row == 3  # header + 2 rows

        # 첫 번째 데이터 row 값 확인
        assert ws.cell(row=2, column=1).value == "post_1"
        assert ws.cell(row=2, column=3).value == "#GC"

    def test_excel_empty_rows(self, tmp_path):
        output = tmp_path / "audit_empty.xlsx"
        write_excel([], output)
        assert output.exists()
        wb = load_workbook(output)
        ws = wb.active
        assert ws.max_row == 1  # header only


# ──────────────────────────────────────────────────────────────
# build_stats 테스트
# ──────────────────────────────────────────────────────────────
class TestBuildStats:
    def test_stats_counts_categories(self):
        rows = [
            {"auto_category": "bag", "parse_status": "ok", "excluded_reason": "", "source_band": "잡화천국22"},
            {"auto_category": "bag", "parse_status": "ok", "excluded_reason": "", "source_band": "잡화천국22"},
            {"auto_category": "top", "parse_status": "ok", "excluded_reason": "", "source_band": "의류천국22"},
            {"auto_category": "etc", "parse_status": "ok", "excluded_reason": "", "source_band": "잡화천국22"},
            {"auto_category": "etc", "parse_status": "parse_error", "excluded_reason": "", "source_band": "잡화천국22"},
        ]
        stats = build_stats(rows)
        assert "총 row 수: 5" in stats
        assert "파싱 성공: 4" in stats
        assert "파싱 실패: 1" in stats
        assert "bag: 2건" in stats
        assert "top: 1건" in stats
        assert "etc: 1건" in stats
        assert "잡화천국22: 4건" in stats
        assert "의류천국22: 1건" in stats

    def test_stats_counts_exclusions(self):
        rows = [
            {"auto_category": "bag", "parse_status": "ok", "excluded_reason": "factory_code:qe", "source_band": "잡화천국22"},
            {"auto_category": "top", "parse_status": "ok", "excluded_reason": "free_size", "source_band": "의류천국22"},
        ]
        stats = build_stats(rows)
        assert "factory_code:qe: 1건" in stats
        assert "free_size: 1건" in stats


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
