"""
[일회성 마이그레이션] reclassify_unclassified.py 단위 테스트.
- DRY RUN 로직 검증 (WC API 호출 없음)
- _infer_source_band / reclassify_product / save_preview_report
"""

from __future__ import annotations

import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

# 스크립트 import 경로 셋업
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from reclassify_unclassified import (
    _infer_source_band,
    fetch_etc_products,
    reclassify_product,
    save_preview_report,
)


# ═══════════════════════════════════════════════════════════════
# 픽스처
# ═══════════════════════════════════════════════════════════════
@pytest.fixture
def sample_db(tmp_path):
    """테스트용 임시 DB 생성 (실제 스키마와 동일)."""
    db_path = tmp_path / "test.db"
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            brand_tag TEXT NOT NULL,
            product_name TEXT NOT NULL,
            set_part TEXT NOT NULL,
            cost_price INTEGER NOT NULL,
            sell_price INTEGER NOT NULL,
            margin_applied INTEGER NOT NULL,
            wc_product_id INTEGER NOT NULL,
            band_key TEXT NOT NULL,
            post_key TEXT NOT NULL,
            category TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    rows = [
        # 잡화 미분류
        ("#MD", "발매트👹", "NULL", 40000, 70000, 30000,
         14862, "97874828", "97874828_94103", "etc"),
        ("#CN", "악세사리 보석함", "NULL", 135000, 165000, 30000,
         14906, "97874828", "97874828_94094", "etc"),
        # 팬티 상품 - 누락된 키워드 "팬티" 때문에 etc
        ("#GC", "벌 패턴 남성팬티 3종세트", "NULL", 41000, 71000, 30000,
         15000, "97874828", "97874828_94200", "etc"),
        # 의류 미분류 - "상하세트" 키워드 매칭될 것
        ("#NK", "[N] 메탈릭 상하세트", "NULL", 43000, 73000, 30000,
         15100, "97874425", "97874425_94300", "etc"),
        # 정상 분류 상품 (재분류 대상 아님)
        ("#LV", "스피디 반둘리에", "NULL", 250000, 300000, 50000,
         15200, "97874828", "97874828_94400", "bag"),
        # 의류 세트 - set_part가 있는 경우
        ("#AZ", "네오테크 후디 셋업 - 상의", "top", 53000, 83000, 30000,
         15300, "97874425", "97874425_94500_top", "etc"),
    ]
    for row in rows:
        c.execute("""
            INSERT INTO products (
                brand_tag, product_name, set_part, cost_price, sell_price,
                margin_applied, wc_product_id, band_key, post_key, category
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, row)
    conn.commit()
    conn.close()
    return str(db_path)


@pytest.fixture
def config():
    """테스트용 최소 설정."""
    return {
        "category_keywords": {
            "bag": ["가방", "백", "반둘리에"],
            "watch": ["시계"],
            "wallet": ["지갑"],
            "shoes": ["스니커즈"],
            "outer": ["자켓", "코트"],
            "top": ["반팔", "티셔츠", "셋업"],
            "bottom": ["팬츠", "데님"],
            "set": ["상하세트", "상하의"],
            "accessory": ["벨트", "모자", "팬티"],  # 팬티 추가 확인
        },
        "wc_categories": {
            "bag": 85, "watch": 86, "accessory": 89,
            "wallet": 83, "shoes": 84, "outer": 78,
            "top": 79, "bottom": 82, "set": 30, "etc": 31,
        },
        "brand_map": {
            "#GC": "구찌", "#NK": "나이키", "#LV": "루이비통",
            "#MD": "MD", "#CN": "채널", "#AZ": "AZ",
        },
    }


# ═══════════════════════════════════════════════════════════════
# 1. _infer_source_band
# ═══════════════════════════════════════════════════════════════
class TestInferSourceBand:
    def test_잡화천국22(self):
        assert _infer_source_band("97874828") == "잡화천국22"

    def test_의류천국22(self):
        assert _infer_source_band("97874425") == "의류천국22"

    def test_unknown_band_key(self):
        assert _infer_source_band("99999999") == "알수없음"

    def test_integer_band_key(self):
        """정수로 들어와도 문자열 변환 후 매칭."""
        assert _infer_source_band(97874828) == "잡화천국22"


# ═══════════════════════════════════════════════════════════════
# 2. fetch_etc_products
# ═══════════════════════════════════════════════════════════════
class TestFetchEtcProducts:
    def test_returns_only_etc(self, sample_db):
        rows = fetch_etc_products(sample_db)
        # 6건 중 etc는 5건 (bag 1건 제외)
        assert len(rows) == 5
        for r in rows:
            assert r["category"] == "etc"

    def test_includes_all_columns(self, sample_db):
        rows = fetch_etc_products(sample_db)
        expected_keys = {
            "id", "brand_tag", "product_name", "set_part",
            "cost_price", "sell_price", "margin_applied",
            "wc_product_id", "band_key", "post_key",
            "category", "created_at"
        }
        assert expected_keys.issubset(set(rows[0].keys()))


# ═══════════════════════════════════════════════════════════════
# 3. reclassify_product (핵심)
# ═══════════════════════════════════════════════════════════════
class TestReclassifyProduct:
    def test_panty_reclassified_to_accessory(self, sample_db, config):
        """팬티 키워드 추가로 accessory 분류 확인."""
        rows = fetch_etc_products(sample_db)
        panty = next(r for r in rows if "팬티" in r["product_name"])
        new_cat = reclassify_product(panty, config)
        assert new_cat == "accessory"

    def test_sanghaeset_reclassified_to_set(self, sample_db, config):
        """상하세트 키워드 + 의류밴드 -> set 분류."""
        rows = fetch_etc_products(sample_db)
        sang = next(r for r in rows if "상하세트" in r["product_name"])
        new_cat = reclassify_product(sang, config)
        assert new_cat == "set"

    def test_set_part_top_forces_set(self, sample_db, config):
        """set_part='top' 은 무조건 set."""
        rows = fetch_etc_products(sample_db)
        set_item = next(r for r in rows if r["set_part"] == "top")
        new_cat = reclassify_product(set_item, config)
        assert new_cat == "set"

    def test_matless_stays_etc(self, sample_db, config):
        """발매트는 키워드 없으므로 여전히 etc."""
        rows = fetch_etc_products(sample_db)
        matless = next(r for r in rows if "발매트" in r["product_name"])
        new_cat = reclassify_product(matless, config)
        assert new_cat == "etc"

    def test_null_set_part_handled(self, sample_db, config):
        """set_part='NULL' 문자열도 None 처리되어야 함."""
        rows = fetch_etc_products(sample_db)
        # 매트나 보석함은 set_part="NULL" 문자열
        matless = next(r for r in rows if "발매트" in r["product_name"])
        assert matless["set_part"] == "NULL"
        new_cat = reclassify_product(matless, config)
        # None 처리되어서 set 카테고리로 가지 않음
        assert new_cat == "etc"


# ═══════════════════════════════════════════════════════════════
# 4. save_preview_report
# ═══════════════════════════════════════════════════════════════
class TestSavePreviewReport:
    def test_creates_xlsx(self, tmp_path, sample_db, config):
        rows = fetch_etc_products(sample_db)
        # 재분류 결과 mock
        items = []
        for r in rows:
            new_cat = reclassify_product(r, config)
            items.append({
                **r,
                "current_category": "etc",
                "new_category": new_cat,
                "source_band": _infer_source_band(r["band_key"]),
            })

        out = tmp_path / "preview.xlsx"
        save_preview_report(items, out)
        assert out.exists()

        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "reclassify_preview"
        # header + 5 rows
        assert ws.max_row == len(items) + 1

    def test_headers_correct(self, tmp_path, sample_db, config):
        rows = fetch_etc_products(sample_db)
        items = [
            {
                **r,
                "current_category": "etc",
                "new_category": "accessory",
                "source_band": "잡화천국22",
            } for r in rows
        ]
        out = tmp_path / "preview.xlsx"
        save_preview_report(items, out)
        wb = load_workbook(out)
        ws = wb.active

        headers = [ws.cell(row=1, column=i).value for i in range(1, 13)]
        expected = [
            "id", "wc_product_id", "brand_tag", "product_name",
            "set_part", "cost_price", "sell_price",
            "current_category", "new_category", "will_change",
            "source_band", "created_at",
        ]
        assert headers == expected

    def test_will_change_flag(self, tmp_path, sample_db, config):
        rows = fetch_etc_products(sample_db)
        items = []
        for r in rows:
            new_cat = reclassify_product(r, config)
            items.append({
                **r,
                "current_category": "etc",
                "new_category": new_cat,
                "source_band": _infer_source_band(r["band_key"]),
            })
        out = tmp_path / "preview.xlsx"
        save_preview_report(items, out)
        wb = load_workbook(out)
        ws = wb.active

        # will_change 컬럼 (10번째)
        change_rows = 0
        for r_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=r_idx, column=10).value
            if val == "CHANGE":
                change_rows += 1
        # 팬티/상하세트/셋업 3건은 CHANGE 여야 함
        assert change_rows == 3


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
