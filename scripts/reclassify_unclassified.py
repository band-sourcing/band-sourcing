#!/usr/bin/env python3
"""
[일회성 마이그레이션] 기존 DB의 etc(미분류) 상품 재분류 DRY RUN.

DB에 raw_content 컬럼이 없으므로 Task 8 토큰 파서 재적용 불가.
대신 현재 product_name을 입력으로 classify_category()만 재실행하여
Task 9 / 키워드 보강 효과를 검증한다.

사용법:
    python scripts/reclassify_unclassified.py           # DRY RUN (엑셀만 생성)
    python scripts/reclassify_unclassified.py --apply   # 실제 DB + WC 업데이트

DRY RUN 결과:
    reports/reclassify_preview_YYYYMMDD_HHMMSS.xlsx
    컬럼: id, wc_product_id, brand_tag, product_name, set_part,
          cost_price, current_category, new_category, will_change
"""

from __future__ import annotations

import argparse
import logging
import os
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# 프로젝트 루트 경로
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import load_config
from src.margin_engine import classify_category

logger = logging.getLogger(__name__)


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


# ═══════════════════════════════════════════════════════════════
# 1. DB 조회
# ═══════════════════════════════════════════════════════════════
def fetch_etc_products(db_path: str) -> list[dict]:
    """category='etc' 인 상품 전체 조회."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT id, brand_tag, product_name, set_part, cost_price, sell_price,
               margin_applied, wc_product_id, band_key, post_key, category, created_at
        FROM products
        WHERE category = 'etc'
        ORDER BY created_at DESC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# 2. 재분류
# ═══════════════════════════════════════════════════════════════
def _infer_source_band(band_key: str) -> str:
    """band_key로부터 source_band 추정.

    97874828 = 잡화천국22
    97874425 = 의류천국22
    """
    mapping = {
        "97874828": "잡화천국22",
        "97874425": "의류천국22",
    }
    return mapping.get(str(band_key), "알수없음")


def reclassify_product(product: dict, config: dict) -> str:
    """현재 product_name + set_part 만으로 재분류 (raw_content 미사용)."""
    source_band = _infer_source_band(product["band_key"])
    # DB의 set_part는 "NULL" 문자열 또는 None 혼재 가능
    set_part = product.get("set_part")
    if set_part in (None, "NULL", "", "nan"):
        set_part = None

    new_cat = classify_category(
        product_name=product["product_name"],
        source_band=source_band,
        category_keywords=config["category_keywords"],
        brand_tag=product["brand_tag"],
        raw_content="",  # DB에 없음 -> 빈 문자열
        set_part=set_part,
    )
    return new_cat


# ═══════════════════════════════════════════════════════════════
# 3. 엑셀 리포트 생성
# ═══════════════════════════════════════════════════════════════
EXCEL_COLUMNS = [
    ("id", 8),
    ("wc_product_id", 12),
    ("brand_tag", 10),
    ("product_name", 45),
    ("set_part", 10),
    ("cost_price", 10),
    ("sell_price", 10),
    ("current_category", 14),
    ("new_category", 14),
    ("will_change", 12),
    ("source_band", 14),
    ("created_at", 20),
]


def save_preview_report(
    items: list[dict],
    output_path: Path,
) -> None:
    """DRY RUN 결과 엑셀 저장."""
    wb = Workbook()
    ws = wb.active
    ws.title = "reclassify_preview"

    # Header 서식
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    change_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 노란색

    for c_idx, (col_name, col_width) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c_idx)].width = col_width

    ws.freeze_panes = "A2"

    # 데이터
    for r_idx, item in enumerate(items, start=2):
        will_change = item["current_category"] != item["new_category"]
        for c_idx, (col_name, _) in enumerate(EXCEL_COLUMNS, start=1):
            value = item.get(col_name, "")
            if col_name == "will_change":
                value = "CHANGE" if will_change else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # 변경될 행은 노란색 배경
            if will_change:
                cell.fill = change_fill

    wb.save(output_path)
    logger.info(f"엑셀 저장 완료: {output_path}")


# ═══════════════════════════════════════════════════════════════
# 4. 실제 변경 (--apply 모드)
# ═══════════════════════════════════════════════════════════════
def apply_changes(
    changes: list[dict],
    db_path: str,
    wc_categories: dict,
    wc_config: dict,
) -> tuple[int, int]:
    """
    실제 WC 카테고리 + DB 업데이트.
    WC 먼저 성공한 후 DB 업데이트 (일관성 우선).

    Returns:
        (성공, 실패) 건수
    """
    from woocommerce import API as WCAPI

    wcapi = WCAPI(
        url=wc_config["url"],
        consumer_key=wc_config["consumer_key"],
        consumer_secret=wc_config["consumer_secret"],
        version="wc/v3",
        timeout=30,
    )

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    success = 0
    failed = 0

    for item in changes:
        wc_id = item["wc_product_id"]
        new_cat = item["new_category"]
        db_id = item["id"]

        new_wc_cat_id = wc_categories.get(new_cat)
        if new_wc_cat_id is None:
            logger.warning(f"WC 카테고리 매핑 없음: {new_cat} (product id={db_id})")
            failed += 1
            continue

        try:
            # 1) WC API로 카테고리 변경
            response = wcapi.put(
                f"products/{wc_id}",
                {"categories": [{"id": new_wc_cat_id}]}
            )
            if response.status_code not in (200, 201):
                logger.error(
                    f"WC 업데이트 실패 (id={db_id}, wc={wc_id}): "
                    f"status={response.status_code}, body={response.text[:200]}"
                )
                failed += 1
                continue

            # 2) DB 업데이트
            cur.execute(
                "UPDATE products SET category = ? WHERE id = ?",
                (new_cat, db_id)
            )
            conn.commit()
            success += 1
            logger.info(
                f"✅ 재분류 [{db_id}] {item['product_name'][:30]} -> "
                f"etc -> {new_cat}"
            )
        except Exception as e:
            logger.error(f"재분류 실패 (id={db_id}): {e}", exc_info=True)
            failed += 1

    conn.close()
    return success, failed


# ═══════════════════════════════════════════════════════════════
# 5. 메인
# ═══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="DB 미분류 상품 재분류 마이그레이션")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="실제 WC + DB 업데이트 (기본은 DRY RUN)"
    )
    parser.add_argument(
        "--db",
        default="data/products.db",
        help="DB 경로 (기본: data/products.db)"
    )
    args = parser.parse_args()

    setup_logging()

    config = load_config()
    db_path = args.db

    if not Path(db_path).exists():
        logger.error(f"DB 파일 없음: {db_path}")
        sys.exit(1)

    # 1) 미분류 상품 조회
    logger.info("=" * 70)
    logger.info(f"[1/4] DB에서 etc 상품 조회: {db_path}")
    etc_products = fetch_etc_products(db_path)
    logger.info(f"    조회된 etc 상품: {len(etc_products)}건")

    if not etc_products:
        logger.info("etc 상품 없음 -> 종료")
        return

    # 2) 각 상품 재분류
    logger.info("[2/4] classify_category 재실행 중...")
    results = []
    stats = {"unchanged": 0, "will_change": 0, "by_new_category": {}}

    for p in etc_products:
        new_cat = reclassify_product(p, config)
        item = {
            **p,
            "current_category": p["category"],
            "new_category": new_cat,
            "source_band": _infer_source_band(p["band_key"]),
        }
        results.append(item)

        if new_cat == "etc":
            stats["unchanged"] += 1
        else:
            stats["will_change"] += 1
            stats["by_new_category"][new_cat] = (
                stats["by_new_category"].get(new_cat, 0) + 1
            )

    # 3) 통계 출력
    logger.info("[3/4] 재분류 결과 통계")
    logger.info(f"    변경 예정: {stats['will_change']}건")
    logger.info(f"    여전히 etc: {stats['unchanged']}건")
    for cat, cnt in sorted(stats["by_new_category"].items(), key=lambda x: -x[1]):
        logger.info(f"      etc -> {cat}: {cnt}건")

    # 4) 엑셀 저장
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = reports_dir / f"reclassify_preview_{ts}.xlsx"
    logger.info(f"[4/4] 엑셀 저장")
    save_preview_report(results, excel_path)

    # 5) 실제 적용 모드
    if args.apply:
        logger.info("=" * 70)
        logger.info("🚨 --apply 모드: 실제 WC + DB 업데이트 시작")
        logger.info(f"총 {stats['will_change']}건 변경 예정")

        # DB 백업
        backup_path = f"{db_path}.backup_{ts}"
        import shutil
        shutil.copy2(db_path, backup_path)
        logger.info(f"✅ DB 백업 완료: {backup_path}")

        # 확인 프롬프트
        print("\n계속 진행하시겠습니까? (yes 입력 시 진행): ", end="")
        confirm = input().strip().lower()
        if confirm != "yes":
            logger.info("사용자가 취소함 -> 종료")
            return

        # 변경 대상만 필터링
        changes = [r for r in results if r["current_category"] != r["new_category"]]

        # WC 설정
        wc_config = {
            "url": os.environ.get("WC_URL", config.get("woocommerce", {}).get("url", "")),
            "consumer_key": os.environ.get(
                "WC_CONSUMER_KEY",
                config.get("woocommerce", {}).get("consumer_key", "")
            ),
            "consumer_secret": os.environ.get(
                "WC_CONSUMER_SECRET",
                config.get("woocommerce", {}).get("consumer_secret", "")
            ),
        }

        if not all(wc_config.values()):
            logger.error("WC 설정 누락 (.env 또는 settings.yaml 확인)")
            sys.exit(1)

        success, failed = apply_changes(
            changes=changes,
            db_path=db_path,
            wc_categories=config["wc_categories"],
            wc_config=wc_config,
        )

        logger.info("=" * 70)
        logger.info(f"✅ 성공: {success}건 / ❌ 실패: {failed}건")
        logger.info(f"💾 DB 백업: {backup_path}")
    else:
        logger.info("=" * 70)
        logger.info("DRY RUN 완료 - 실제 변경 없음")
        logger.info(f"엑셀 검토 후 --apply 로 실행하세요:")
        logger.info(f"    python {sys.argv[0]} --apply")


if __name__ == "__main__":
    main()
