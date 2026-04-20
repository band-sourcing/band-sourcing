#!/usr/bin/env python3
"""
[Audit] 밴드 크롤링 + 현재 분류 결과 감사 스크립트.

목적:
  잡화천국22 / 의류천국22 두 밴드의 cutoff_date 이후 전체 게시글을 크롤링하여
  현재 classify_category() 분류 결과를 엑셀로 출력. 사용자가 엑셀에서
  correct_category 컬럼을 채우면 Phase 2(오분류 분석)로 이어진다.

실행 (서버):
  cd /opt/band-sourcing
  source venv/bin/activate
  pip install openpyxl
  python3 scripts/audit_crawl_and_classify.py

실행 (로컬):
  python scripts/audit_crawl_and_classify.py

출력:
  outputs/audit_YYYYMMDD_HHMMSS.xlsx
  outputs/audit_YYYYMMDD_HHMMSS_stats.txt
"""

from __future__ import annotations

import logging
import os
import sys
from dataclasses import asdict
from datetime import datetime
from pathlib import Path

# 프로젝트 루트를 sys.path에 추가
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.band_scraper import BandScraper
from src.config import load_config
from src.content_parser import ParseError, parse_post
from src.exclusion_filter import should_exclude
from src.margin_engine import classify_category


# ──────────────────────────────────────────────────────────────
# 로깅 설정
# ──────────────────────────────────────────────────────────────
def setup_logging(log_dir: Path) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger("audit")


# ──────────────────────────────────────────────────────────────
# 엑셀 헤더 정의
# ──────────────────────────────────────────────────────────────
EXCEL_COLUMNS = [
    ("post_key", 28),
    ("source_band", 14),
    ("brand_tag", 10),
    ("brand_name_kr", 16),
    ("product_name", 40),
    ("price_code", 10),
    ("cost_price", 12),
    ("sizes", 24),
    ("auto_category", 14),
    ("excluded_reason", 16),   # 제외 필터 사유 (factory_code / free_size / "")
    ("parse_status", 14),       # ok / parse_error / no_content
    ("correct_category", 14),  # Phase 2에서 사용자가 수동 기입
    ("txtBody_raw", 120),
]

# 카테고리별 색상 (시각적 검토 편의)
CATEGORY_FILL = {
    "bag":       PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid"),
    "watch":     PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid"),
    "wallet":    PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),
    "shoes":     PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    "outer":     PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid"),
    "top":       PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),
    "bottom":    PatternFill(start_color="F0E68C", end_color="F0E68C", fill_type="solid"),
    "accessory": PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"),
    "etc":       PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),  # 미분류 빨강
}


# ──────────────────────────────────────────────────────────────
# Row 생성 helper
# ──────────────────────────────────────────────────────────────
def build_row_from_product(
    product,
    post_key: str,
    source_band: str,
    brand_map: dict,
    category_keywords: dict,
    exclusion_config: dict,
    txtbody_raw: str,
) -> dict:
    """파싱 성공 상품에서 audit row 생성."""
    # post_key에 set_part suffix (세트 상품 대응)
    row_post_key = post_key
    if product.set_part:
        row_post_key = f"{post_key}_{product.set_part}"

    brand_name_kr = brand_map.get(product.brand_tag, "")

    auto_category = classify_category(
        product.product_name,
        product.source_band,
        category_keywords,
        brand_tag=product.brand_tag,
        raw_content=txtbody_raw,
        set_part=product.set_part,
    )

    # 제외 필터 판정 (원본 dict를 건드리지 않고 사유만 기록)
    excluded_reason = ""
    fc_config = exclusion_config.get("factory_codes", {})
    if fc_config.get("enabled", False):
        codes = [c.lower() for c in fc_config.get("codes", [])]
        sc = (product.season_code or "").strip().lower()
        if sc in codes:
            excluded_reason = f"factory_code:{product.season_code}"

    if not excluded_reason:
        fs_config = exclusion_config.get("free_size", {})
        if fs_config.get("enabled", False):
            target_bands = fs_config.get("target_bands", [])
            if product.source_band in target_bands:
                sizes_upper = [str(s).upper() for s in product.sizes]
                if any("FREE" in s or s == "F" for s in sizes_upper):
                    excluded_reason = "free_size"

    return {
        "post_key": row_post_key,
        "source_band": source_band,
        "brand_tag": product.brand_tag,
        "brand_name_kr": brand_name_kr,
        "product_name": product.product_name,
        "price_code": product.season_code,
        "cost_price": product.cost_price,
        "sizes": ", ".join(product.sizes) if product.sizes else "",
        "auto_category": auto_category,
        "excluded_reason": excluded_reason,
        "parse_status": "ok",
        "correct_category": "",
        "txtBody_raw": txtbody_raw,
    }


def build_row_from_parse_error(
    post_key: str,
    source_band: str,
    txtbody_raw: str,
    error_msg: str,
) -> dict:
    """파싱 실패 게시글용 audit row (수동 검토 대상)."""
    return {
        "post_key": post_key,
        "source_band": source_band,
        "brand_tag": "",
        "brand_name_kr": "",
        "product_name": f"[PARSE_ERROR] {error_msg}",
        "price_code": "",
        "cost_price": 0,
        "sizes": "",
        "auto_category": "etc",
        "excluded_reason": "",
        "parse_status": "parse_error",
        "correct_category": "",
        "txtBody_raw": txtbody_raw,
    }


# ──────────────────────────────────────────────────────────────
# 엑셀 쓰기
# ──────────────────────────────────────────────────────────────
def write_excel(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "audit"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, (col_name, col_width) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = "A2"

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (col_name, _) in enumerate(EXCEL_COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # txtBody_raw 컬럼은 줄바꿈 래핑
            if col_name == "txtBody_raw":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # auto_category 색상
        auto_cat = row.get("auto_category", "")
        fill = CATEGORY_FILL.get(auto_cat)
        if fill:
            auto_col_idx = [c[0] for c in EXCEL_COLUMNS].index("auto_category") + 1
            ws.cell(row=r_idx, column=auto_col_idx).fill = fill

    # Row height (txtBody_raw 가독성)
    for r_idx in range(2, len(rows) + 2):
        ws.row_dimensions[r_idx].height = 60

    wb.save(output_path)


# ──────────────────────────────────────────────────────────────
# 통계 집계
# ──────────────────────────────────────────────────────────────
def build_stats(rows: list[dict]) -> str:
    from collections import Counter

    total = len(rows)
    cat_counter = Counter(r["auto_category"] for r in rows if r["parse_status"] == "ok")
    excluded_counter = Counter(r["excluded_reason"] for r in rows if r["excluded_reason"])
    parse_error_count = sum(1 for r in rows if r["parse_status"] == "parse_error")
    band_counter = Counter(r["source_band"] for r in rows)

    lines = []
    lines.append("=" * 60)
    lines.append("Audit 분류 통계")
    lines.append("=" * 60)
    lines.append(f"총 row 수: {total}")
    lines.append(f"파싱 성공: {total - parse_error_count}")
    lines.append(f"파싱 실패: {parse_error_count}")
    lines.append("")
    lines.append("[밴드별 분포]")
    for band, cnt in band_counter.most_common():
        lines.append(f"  {band}: {cnt}건")
    lines.append("")
    lines.append("[auto_category 분포 (파싱 성공 기준)]")
    ok_total = sum(cat_counter.values())
    priority = ["bag", "watch", "wallet", "shoes", "outer", "top", "bottom", "accessory", "etc"]
    for cat in priority:
        cnt = cat_counter.get(cat, 0)
        pct = (cnt / ok_total * 100) if ok_total else 0
        marker = " ⚠️  미분류" if cat == "etc" and cnt > 0 else ""
        lines.append(f"  {cat}: {cnt}건 ({pct:.1f}%){marker}")
    lines.append("")
    lines.append("[제외 필터 적용 분포]")
    if excluded_counter:
        for reason, cnt in excluded_counter.most_common():
            lines.append(f"  {reason}: {cnt}건")
    else:
        lines.append("  (없음)")
    lines.append("")
    lines.append("=" * 60)
    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────────
def main():
    log_dir = PROJECT_ROOT / "logs"
    logger = setup_logging(log_dir)

    logger.info("=" * 60)
    logger.info("Band Audit 스크립트 시작")
    logger.info("=" * 60)

    config = load_config()
    brand_map = config["brand_map"]
    category_keywords = config["category_keywords"]
    exclusion_config = config.get("exclusion", {})

    # 출력 디렉토리
    output_dir = PROJECT_ROOT / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"audit_{timestamp}.xlsx"
    stats_path = output_dir / f"audit_{timestamp}_stats.txt"

    # 크롤링
    logger.info("=== Step 1: Band 크롤링 시작 ===")
    scraper = BandScraper(
        naver_id=config["band"]["naver_id"],
        naver_pw=config["band"]["naver_pw"],
        cutoff_date=config["band"]["cutoff_date"],
        session_path="data/band_session.json",
    )

    try:
        if config["band"].get("band_keys"):
            band_keys = config["band"]["band_keys"]
        else:
            band_keys = scraper.get_band_keys(config["band"]["target_bands"])
        logger.info(f"대상 밴드: {list(band_keys.keys())}")

        all_posts = []
        for band_name, band_key in band_keys.items():
            posts = scraper.fetch_all_posts(band_key)
            for p in posts:
                p["_source_band"] = band_name
            all_posts.extend(posts)
            logger.info(f"  {band_name}: {len(posts)}개 수집")

        logger.info(f"총 {len(all_posts)}개 게시글 수집 완료")
    finally:
        scraper.close()

    # 파싱 + 분류
    logger.info("=== Step 2: 파싱 및 현재 분류 로직 적용 ===")
    rows: list[dict] = []
    parse_error_count = 0

    for post in all_posts:
        post_key = post["post_key"]
        source_band = post.get("_source_band", "")
        content = post.get("content", "") or ""

        if not content.strip():
            rows.append({
                "post_key": post_key,
                "source_band": source_band,
                "brand_tag": "",
                "brand_name_kr": "",
                "product_name": "[NO_CONTENT]",
                "price_code": "",
                "cost_price": 0,
                "sizes": "",
                "auto_category": "etc",
                "excluded_reason": "",
                "parse_status": "no_content",
                "correct_category": "",
                "txtBody_raw": "",
            })
            continue

        try:
            products = parse_post(
                content=content,
                brand_map=brand_map,
                source_band=source_band,
            )
            for product in products:
                row = build_row_from_product(
                    product=product,
                    post_key=post_key,
                    source_band=source_band,
                    brand_map=brand_map,
                    category_keywords=category_keywords,
                    exclusion_config=exclusion_config,
                    txtbody_raw=content,
                )
                rows.append(row)
        except ParseError as e:
            parse_error_count += 1
            logger.warning(f"파싱 실패 (post_key={post_key}): {e}")
            rows.append(build_row_from_parse_error(
                post_key=post_key,
                source_band=source_band,
                txtbody_raw=content,
                error_msg=str(e),
            ))
        except Exception as e:
            logger.error(f"예외 발생 (post_key={post_key}): {e}", exc_info=True)
            rows.append(build_row_from_parse_error(
                post_key=post_key,
                source_band=source_band,
                txtbody_raw=content,
                error_msg=f"UNEXPECTED: {type(e).__name__}: {e}",
            ))

    logger.info(f"총 audit row: {len(rows)}개 (파싱 실패 {parse_error_count}건 포함)")

    # 엑셀 출력
    logger.info("=== Step 3: 엑셀 출력 ===")
    write_excel(rows, xlsx_path)
    logger.info(f"엑셀 저장: {xlsx_path}")

    # 통계 출력
    logger.info("=== Step 4: 분류 통계 ===")
    stats_text = build_stats(rows)
    print("\n" + stats_text + "\n")
    with stats_path.open("w", encoding="utf-8") as f:
        f.write(stats_text)
    logger.info(f"통계 저장: {stats_path}")

    logger.info("=" * 60)
    logger.info("Audit 완료")
    logger.info(f"  엑셀: {xlsx_path}")
    logger.info(f"  통계: {stats_path}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
