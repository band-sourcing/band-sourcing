#!/usr/bin/env python3
"""
[일회성 마이그레이션] DB의 etc(미분류) 상품을 개별 게시글 재크롤링 후 재분류.

기존 reclassify_unclassified.py 의 확장판:
- reclassify_unclassified.py: DB의 product_name 만 재사용 (망가진 상품명은 고치지 못함)
- recrawl_and_reclassify.py (이 파일): post_key 로 원본 게시글 재방문
                                        -> 토큰 파서로 상품명 재추출
                                        -> 재분류 + WC 업데이트

사용법:
    # DRY RUN (엑셀 리포트 생성만)
    python scripts/recrawl_and_reclassify.py

    # 실제 DB + WC 업데이트
    python scripts/recrawl_and_reclassify.py --apply

DRY RUN 결과:
    reports/recrawl_preview_YYYYMMDD_HHMMSS.xlsx
"""

from __future__ import annotations

import argparse
import logging
import os
import shutil
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.band_scraper import BandScraper
from src.config import load_config
from src.content_parser import parse_post, ParseError
from src.margin_engine import calculate_sell_price, classify_category

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# 유틸
# ═══════════════════════════════════════════════════════════════
def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def _infer_source_band(band_key: str) -> str:
    mapping = {
        "97874828": "잡화천국22",
        "97874425": "의류천국22",
    }
    return mapping.get(str(band_key), "알수없음")


def _build_post_url(band_key: str, post_id: str) -> str:
    """post_key에서 band detail URL 생성.

    post_key 형식: {band_key}_{post_id}
    URL: https://band.us/band/{band_key}/post/{post_id}
    """
    return f"https://band.us/band/{band_key}/post/{post_id}"


# ═══════════════════════════════════════════════════════════════
# DB 조회
# ═══════════════════════════════════════════════════════════════
def fetch_etc_products(db_path: str) -> list[dict]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("""
        SELECT id, brand_tag, product_name, set_part, cost_price, sell_price,
               margin_applied, wc_product_id, band_key, post_key, category, created_at
        FROM products
        WHERE category = 'etc'
        ORDER BY created_at DESC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# 재크롤링 + 재파싱
# ═══════════════════════════════════════════════════════════════
def recrawl_post_content(scraper: BandScraper, band_key: str, post_id: str) -> str | None:
    """
    단일 게시글을 재방문하여 txtBody 재추출.
    실패 시 None 반환.
    """
    post_url = _build_post_url(band_key, post_id)

    try:
        scraper._safe_goto(post_url)
        time.sleep(1)
    except Exception as e:
        logger.warning(f"  goto 실패: {post_url} -> {e}")
        return None

    detail_selectors = [
        'p.txtBody',
        '.postText',
        '._postText',
        '[class*="postText"]',
        '[class*="txtBody"]',
        '.dPostBody',
        '._postBody',
    ]

    for sel in detail_selectors:
        try:
            text_el = scraper._page.locator(sel).first
            if text_el.count() > 0:
                text = text_el.inner_text().strip()
                if text and len(text) > 3:
                    return text
        except Exception as e:
            err_msg = str(e).lower()
            if "execution context" in err_msg or "destroyed" in err_msg:
                try:
                    scraper._wait_for_stable_context()
                except Exception:
                    pass
                break
            continue

    return None


def reparse_product(
    raw_content: str,
    source_band: str,
    brand_tag_hint: str,
    set_part_hint: str | None,
    config: dict,
) -> dict | None:
    """
    raw_content 로부터 parse_post + classify_category 실행.

    Returns:
        {
          "product_name": str,
          "brand_tag": str,
          "cost_price": int,
          "category": str,
          "set_part": str | None,
        } 또는 None (파싱 실패)
    """
    try:
        products = parse_post(raw_content, config["brand_map"], source_band)
    except ParseError as e:
        logger.debug(f"  parse_post 실패: {e}")
        return None

    if not products:
        return None

    # 원본 DB 행의 set_part 에 맞는 product 선택
    # set_part 힌트가 있으면 그에 맞는 것 없으면 첫 번째
    target = None
    if set_part_hint in ("top", "bottom"):
        for p in products:
            if p.set_part == set_part_hint:
                target = p
                break
    if target is None:
        target = products[0]

    category = classify_category(
        target.product_name,
        source_band,
        config["category_keywords"],
        brand_tag=target.brand_tag,
        raw_content=raw_content,
        set_part=target.set_part,
    )

    return {
        "product_name": target.product_name,
        "brand_tag": target.brand_tag,
        "cost_price": target.cost_price,
        "category": category,
        "set_part": target.set_part,
    }


# ═══════════════════════════════════════════════════════════════
# 엑셀 리포트
# ═══════════════════════════════════════════════════════════════
EXCEL_COLUMNS = [
    ("id", 8),
    ("wc_product_id", 12),
    ("band_key", 12),
    ("post_id", 10),
    ("brand_tag", 10),
    ("set_part", 10),
    ("old_product_name", 40),
    ("new_product_name", 40),
    ("old_cost_price", 10),
    ("new_cost_price", 10),
    ("old_category", 14),
    ("new_category", 14),
    ("change_type", 15),
    ("source_band", 14),
    ("created_at", 20),
]


def save_preview_report(items: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "recrawl_preview"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    change_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for c_idx, (col_name, col_width) in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c_idx)].width = col_width

    ws.freeze_panes = "A2"

    for r_idx, item in enumerate(items, start=2):
        ct = item.get("change_type", "")
        for c_idx, (col_name, _) in enumerate(EXCEL_COLUMNS, start=1):
            value = item.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if ct == "CATEGORY_CHANGE" or ct == "NAME_CHANGE" or ct == "BOTH_CHANGE":
                cell.fill = change_fill
            elif ct == "RECRAWL_FAILED" or ct == "PARSE_FAILED":
                cell.fill = error_fill

    wb.save(output_path)
    logger.info(f"엑셀 저장 완료: {output_path}")


# ═══════════════════════════════════════════════════════════════
# WC + DB 업데이트
# ═══════════════════════════════════════════════════════════════
def apply_changes(
    changes: list[dict],
    db_path: str,
    wc_categories: dict,
    wc_config: dict,
    update_name: bool,
    update_price: bool,
) -> tuple[int, int]:
    """
    WC 카테고리 + 이름 + 가격 업데이트 후 DB 업데이트.
    """
    from woocommerce import API as WCAPI

    wcapi = WCAPI(
        url=wc_config["url"],
        consumer_key=wc_config["consumer_key"],
        consumer_secret=wc_config["consumer_secret"],
        version="wc/v3",
        timeout=30,
    )

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    success = 0
    failed = 0

    for item in changes:
        wc_id = item["wc_product_id"]
        new_cat = item["new_category"]
        db_id = item["id"]

        new_wc_cat_id = wc_categories.get(new_cat)
        if new_wc_cat_id is None:
            logger.warning(f"WC 카테고리 매핑 없음: {new_cat} (id={db_id})")
            failed += 1
            continue

        wc_payload: dict = {"categories": [{"id": new_wc_cat_id}]}
        if update_name and item.get("new_product_name"):
            wc_payload["name"] = item["new_product_name"]
        # 가격 변경은 margin 재계산 후에만
        if update_price and item.get("new_sell_price"):
            wc_payload["regular_price"] = str(item["new_sell_price"])

        try:
            response = wcapi.put(f"products/{wc_id}", wc_payload)
            if response.status_code not in (200, 201):
                logger.error(
                    f"WC 업데이트 실패 (id={db_id}, wc={wc_id}): "
                    f"status={response.status_code}, body={response.text[:200]}"
                )
                failed += 1
                continue

            # DB 업데이트
            db_updates = ["category = ?"]
            db_values = [new_cat]
            if update_name and item.get("new_product_name"):
                db_updates.append("product_name = ?")
                db_values.append(item["new_product_name"])
            if update_price:
                if item.get("new_cost_price"):
                    db_updates.append("cost_price = ?")
                    db_values.append(item["new_cost_price"])
                if item.get("new_sell_price"):
                    db_updates.append("sell_price = ?")
                    db_values.append(item["new_sell_price"])
                if item.get("new_margin"):
                    db_updates.append("margin_applied = ?")
                    db_values.append(item["new_margin"])

            db_values.append(db_id)
            cur.execute(
                f"UPDATE products SET {', '.join(db_updates)} WHERE id = ?",
                db_values
            )
            conn.commit()
            success += 1
            logger.info(
                f"✅ [{db_id}] {item.get('old_product_name','')[:25]} -> "
                f"{item.get('new_product_name','')[:25]} [{item['old_category']} -> {new_cat}]"
            )
        except Exception as e:
            logger.error(f"업데이트 실패 (id={db_id}): {e}", exc_info=True)
            failed += 1

    conn.close()
    return success, failed


# ═══════════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="etc 상품 재크롤링 + 재분류")
    parser.add_argument("--apply", action="store_true", help="실제 WC+DB 업데이트")
    parser.add_argument("--db", default="data/products.db")
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="처리 건수 제한 (0=전체, 테스트용)"
    )
    args = parser.parse_args()

    setup_logging()
    config = load_config()
    db_path = args.db

    if not Path(db_path).exists():
        logger.error(f"DB 파일 없음: {db_path}")
        sys.exit(1)

    # 1) DB에서 etc 조회
    logger.info("=" * 70)
    logger.info(f"[1/5] DB에서 etc 상품 조회")
    etc_products = fetch_etc_products(db_path)
    logger.info(f"    조회: {len(etc_products)}건")
    if args.limit > 0:
        etc_products = etc_products[:args.limit]
        logger.info(f"    --limit 적용: {len(etc_products)}건만 처리")

    if not etc_products:
        logger.info("etc 0건 -> 종료")
        return

    # 2) Scraper 초기화
    logger.info(f"[2/5] BandScraper 초기화")
    scraper = BandScraper(
        naver_id=config["band"]["naver_id"],
        naver_pw=config["band"]["naver_pw"],
        cutoff_date=config["band"]["cutoff_date"],
        session_path="data/band_session.json",
    )
    try:
        scraper.ensure_logged_in()
    except Exception as e:
        logger.error(f"❌ 밴드 로그인 실패: {e}")
        scraper.close()
        sys.exit(1)

    # 3) 재크롤링 + 재파싱
    logger.info(f"[3/5] 재크롤링 + 재파싱 시작 ({len(etc_products)}건)")
    results = []
    stats = {
        "total": len(etc_products),
        "category_change": 0,
        "name_change": 0,
        "both_change": 0,
        "no_change": 0,
        "recrawl_failed": 0,
        "parse_failed": 0,
    }

    try:
        for idx, p in enumerate(etc_products, 1):
            if idx % 20 == 0:
                logger.info(f"  진행: {idx}/{len(etc_products)}")

            band_key = str(p["band_key"])
            # post_key 형식: {band_key}_{post_id}
            post_key_full = p["post_key"]
            if post_key_full.startswith(f"{band_key}_"):
                post_id = post_key_full[len(band_key) + 1:]
            else:
                post_id = post_key_full

            source_band = _infer_source_band(band_key)
            old_set_part = p.get("set_part")
            if old_set_part in (None, "NULL", "", "nan"):
                old_set_part = None

            # 재크롤링
            raw_content = recrawl_post_content(scraper, band_key, post_id)

            item = {
                "id": p["id"],
                "wc_product_id": p["wc_product_id"],
                "band_key": band_key,
                "post_id": post_id,
                "brand_tag": p["brand_tag"],
                "set_part": old_set_part or "",
                "old_product_name": p["product_name"],
                "old_cost_price": p["cost_price"],
                "old_category": p["category"],
                "source_band": source_band,
                "created_at": str(p["created_at"]),
            }

            if not raw_content:
                item["change_type"] = "RECRAWL_FAILED"
                item["new_product_name"] = ""
                item["new_cost_price"] = ""
                item["new_category"] = ""
                stats["recrawl_failed"] += 1
                results.append(item)
                continue

            # 재파싱
            reparsed = reparse_product(
                raw_content=raw_content,
                source_band=source_band,
                brand_tag_hint=p["brand_tag"],
                set_part_hint=old_set_part,
                config=config,
            )
            if reparsed is None:
                item["change_type"] = "PARSE_FAILED"
                item["new_product_name"] = ""
                item["new_cost_price"] = ""
                item["new_category"] = ""
                stats["parse_failed"] += 1
                results.append(item)
                continue

            # 새 가격 + 판매가
            new_sell_price, new_margin = calculate_sell_price(
                reparsed["cost_price"],
                reparsed["category"],
                config["margin"],
            )
            item["new_product_name"] = reparsed["product_name"]
            item["new_cost_price"] = reparsed["cost_price"]
            item["new_sell_price"] = new_sell_price
            item["new_margin"] = new_margin
            item["new_category"] = reparsed["category"]

            # 변경 유형 판정
            name_changed = item["old_product_name"] != reparsed["product_name"]
            cat_changed = item["old_category"] != reparsed["category"]

            if name_changed and cat_changed:
                item["change_type"] = "BOTH_CHANGE"
                stats["both_change"] += 1
            elif cat_changed:
                item["change_type"] = "CATEGORY_CHANGE"
                stats["category_change"] += 1
            elif name_changed:
                item["change_type"] = "NAME_CHANGE"
                stats["name_change"] += 1
            else:
                item["change_type"] = "NO_CHANGE"
                stats["no_change"] += 1

            results.append(item)
    finally:
        scraper.close()

    # 4) 통계
    logger.info("=" * 70)
    logger.info("[4/5] 재크롤링 결과 통계")
    logger.info(f"    총 처리: {stats['total']}")
    logger.info(f"    카테고리 + 이름 모두 변경: {stats['both_change']}")
    logger.info(f"    카테고리만 변경: {stats['category_change']}")
    logger.info(f"    이름만 변경: {stats['name_change']}")
    logger.info(f"    변경 없음: {stats['no_change']}")
    logger.info(f"    재크롤링 실패: {stats['recrawl_failed']}")
    logger.info(f"    파싱 실패: {stats['parse_failed']}")

    # 5) 엑셀 저장
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = reports_dir / f"recrawl_preview_{ts}.xlsx"
    logger.info(f"[5/5] 엑셀 저장")
    save_preview_report(results, excel_path)

    # --apply
    if args.apply:
        logger.info("=" * 70)
        logger.info("🚨 --apply: 실제 WC + DB 업데이트 시작")
        changes = [
            r for r in results
            if r["change_type"] in ("CATEGORY_CHANGE", "NAME_CHANGE", "BOTH_CHANGE")
        ]
        logger.info(f"변경 대상: {len(changes)}건")

        backup_path = f"{db_path}.backup_{ts}"
        shutil.copy2(db_path, backup_path)
        logger.info(f"✅ DB 백업 완료: {backup_path}")

        print("\n계속 진행하시겠습니까? (yes 입력 시 진행): ", end="")
        confirm = input().strip().lower()
        if confirm != "yes":
            logger.info("취소 -> 종료")
            return

        wc_config = {
            "url": os.environ.get("WC_URL", config.get("woocommerce", {}).get("url", "")),
            "consumer_key": os.environ.get(
                "WC_CONSUMER_KEY",
                config.get("woocommerce", {}).get("consumer_key", "")
            ),
            "consumer_secret": os.environ.get(
                "WC_CONSUMER_SECRET",
                config.get("woocommerce", {}).get("consumer_secret", "")
            ),
        }

        if not all(wc_config.values()):
            logger.error("WC 설정 누락 (.env 또는 settings.yaml 확인)")
            sys.exit(1)

        success, failed = apply_changes(
            changes=changes,
            db_path=db_path,
            wc_categories=config["wc_categories"],
            wc_config=wc_config,
            update_name=True,
            update_price=True,
        )

        logger.info("=" * 70)
        logger.info(f"✅ 성공: {success}건 / ❌ 실패: {failed}건")
        logger.info(f"💾 DB 백업: {backup_path}")
    else:
        logger.info("=" * 70)
        logger.info("DRY RUN 완료 - 실제 변경 없음")
        logger.info(f"엑셀 검토 후 --apply 실행:")
        logger.info(f"    python {sys.argv[0]} --apply")


if __name__ == "__main__":
    main()
